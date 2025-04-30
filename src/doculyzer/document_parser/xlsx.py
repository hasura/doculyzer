"""
XLSX document parser module for the document pointer system.

This module parses Excel (XLSX) files into structured elements.
"""

import hashlib
import json
import logging
import os
import uuid
from typing import Dict, Any, List, Optional, Union

try:
    # noinspection PyUnresolvedReferences
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Install with 'pip install openpyxl' to use XLSX parser")

from .base import DocumentParser

logger = logging.getLogger(__name__)


class XlsxParser(DocumentParser):
    """Parser for Excel (XLSX) documents."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize the XLSX parser."""
        super().__init__(config)

        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX parsing")

        # Configuration options
        self.config = config or {}
        self.extract_hidden_sheets = self.config.get("extract_hidden_sheets", False)
        self.extract_formulas = self.config.get("extract_formulas", True)
        self.extract_comments = self.config.get("extract_comments", True)
        self.extract_charts = self.config.get("extract_charts", False)
        self.extract_images = self.config.get("extract_images", False)
        self.max_rows = self.config.get("max_rows", 1000)  # Limit for large spreadsheets
        self.max_cols = self.config.get("max_cols", 100)  # Limit for very wide sheets
        self.temp_dir = self.config.get("temp_dir", os.path.join(os.path.dirname(__file__), 'temp'))
        self.max_content_preview = self.config.get("max_content_preview", 100)

        # Data table detection options
        self.detect_tables = self.config.get("detect_tables", True)  # Whether to detect data tables
        self.min_table_rows = self.config.get("min_table_rows", 2)  # Minimum rows for table detection
        self.min_table_cols = self.config.get("min_table_cols", 2)  # Minimum columns for table detection

    def parse(self, doc_content: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse an XLSX document into structured elements.

        Args:
            doc_content: Document content and metadata

        Returns:
            Dictionary with document metadata, elements, relationships, and extracted links
        """
        # Extract metadata from doc_content
        source_id = doc_content["id"]
        metadata = doc_content.get("metadata", {}).copy()  # Make a copy to avoid modifying original

        # Check if we have binary content or a path to a file
        binary_path = doc_content.get("binary_path")
        if not binary_path:
            # If we have binary content but no path, we need to save it to a temp file
            if not os.path.exists(self.temp_dir):
                os.makedirs(self.temp_dir, exist_ok=True)

            binary_content = doc_content.get("content", b"")
            if isinstance(binary_content, str):
                logger.warning("Expected binary content for XLSX but got string. Attempting to process anyway.")

            temp_file_path = os.path.join(self.temp_dir, f"temp_{uuid.uuid4().hex}.xlsx")
            with open(temp_file_path, 'wb') as f:
                if isinstance(binary_content, str):
                    f.write(binary_content.encode('utf-8'))
                else:
                    f.write(binary_content)

            binary_path = temp_file_path
            logger.debug(f"Saved binary content to temporary file: {binary_path}")

        # Generate document ID if not present
        doc_id = metadata.get("doc_id", self._generate_id("doc_"))

        # Load XLSX document
        try:
            # Use read_only mode for better performance with large files
            workbook = openpyxl.load_workbook(binary_path, read_only=True, data_only=not self.extract_formulas)
        except Exception as e:
            logger.error(f"Error loading XLSX document: {str(e)}")
            raise

        # Create document record with metadata
        document = {
            "doc_id": doc_id,
            "doc_type": "xlsx",
            "source": source_id,
            "metadata": self._extract_document_metadata(workbook, metadata),
            "content_hash": doc_content.get("content_hash", "")
        }

        # Create root element
        elements = [self._create_root_element(doc_id, source_id)]
        root_id = elements[0]["element_id"]

        # Parse document elements
        sheet_elements = self._parse_workbook(workbook, doc_id, root_id, source_id)
        elements.extend(sheet_elements)

        # Extract links from the document using the new helper method
        links = self._extract_workbook_links(sheet_elements)

        # Clean up temporary file if needed
        if binary_path != doc_content.get("binary_path") and os.path.exists(binary_path):
            try:
                os.remove(binary_path)
                logger.debug(f"Deleted temporary file: {binary_path}")
            except Exception as e:
                logger.warning(f"Failed to delete temporary file {binary_path}: {str(e)}")

        # Close workbook
        workbook.close()

        # Return the parsed document with extracted links
        return {
            "document": document,
            "elements": elements,
            "links": links,
            "relationships": []
        }

    def _resolve_element_content(self, location_data: Dict[str, Any],
                                 source_content: Optional[Union[str, bytes]]) -> str:
        """
        Resolve content for specific XLSX element types.

        Args:
            location_data: Content location data
            source_content: Optional pre-loaded source content

        Returns:
            Resolved content string
        """
        source = location_data.get("source", "")
        element_type = location_data.get("type", "")
        sheet_name = location_data.get("sheet_name", "")

        # Load the document if source content is not provided
        wb = None
        temp_file = None
        try:
            if source_content is None:
                # Check if source is a file path
                if os.path.exists(source):
                    try:
                        wb = openpyxl.load_workbook(source, read_only=True, data_only=not self.extract_formulas)
                    except Exception as e:
                        raise ValueError(f"Error loading XLSX document: {str(e)}")
                else:
                    raise ValueError(f"Source file not found: {source}")
            else:
                # Save content to a temporary file
                if not os.path.exists(self.temp_dir):
                    os.makedirs(self.temp_dir, exist_ok=True)

                import uuid
                temp_file = os.path.join(self.temp_dir, f"temp_{uuid.uuid4().hex}.xlsx")
                with open(temp_file, 'wb') as f:
                    if isinstance(source_content, str):
                        f.write(source_content.encode('utf-8'))
                    else:
                        f.write(source_content)

                # Load the document
                try:
                    wb = openpyxl.load_workbook(temp_file, read_only=True, data_only=not self.extract_formulas)
                except Exception as e:
                    raise ValueError(f"Error loading XLSX document: {str(e)}")

            # Handle different element types
            if element_type == "workbook":
                # Return information about the workbook
                sheet_names = wb.sheetnames
                active_sheet = wb.active.title if hasattr(wb, 'active') and wb.active else None
                return f"Workbook with sheets: {', '.join(sheet_names)}. Active sheet: {active_sheet or 'None'}"

            # Check if sheet exists
            if sheet_name and sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            # Get the specified sheet
            if sheet_name:
                sheet = wb[sheet_name]
            else:
                # Use active sheet if no specific sheet name provided
                sheet = wb.active

            if element_type == "sheet":
                # Return information about the sheet
                max_row = min(sheet.max_row or 0, self.max_rows)
                max_col = min(sheet.max_column or 0, self.max_cols)
                return f"Sheet '{sheet.title}' with {max_row} rows and {max_col} columns"

            elif element_type == "table_row":
                # Extract row by index
                row = location_data.get("row", 0)

                if row <= 0 or row > min(sheet.max_row or 0, self.max_rows):
                    return f"Row {row} is out of range"

                row_values = []
                for col in range(1, min(sheet.max_column + 1, self.max_cols + 1)):
                    cell = sheet.cell(row=row, column=col)
                    row_values.append(str(cell.value) if cell.value is not None else "")

                return "\t".join(row_values)

            elif element_type == "table_cell":
                # Extract cell by reference
                cell_ref = location_data.get("cell", "")

                if cell_ref:
                    # Direct cell reference (e.g., "A1")
                    try:
                        cell = sheet[cell_ref]
                        return str(cell.value) if cell.value is not None else ""
                    except Exception as e:
                        return f"Error accessing cell {cell_ref}: {str(e)}"
                else:
                    # Row/column coordinates
                    row = location_data.get("row", 0)
                    col = location_data.get("col", 0)

                    if row <= 0 or col <= 0 or row > min(sheet.max_row or 0, self.max_rows) or col > min(
                            sheet.max_column or 0, self.max_cols):
                        return f"Cell at row {row}, column {col} is out of range"

                    cell = sheet.cell(row=row, column=col)
                    return str(cell.value) if cell.value is not None else ""

            elif element_type == "data_table":
                # Extract a range of cells forming a table
                range_str = location_data.get("range", "")

                if not range_str:
                    return "No range specified for data table"

                try:
                    # Get all cells in the range
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(range_str)

                    # Extract table data
                    table_data = []
                    for row in range(min_row, max_row + 1):
                        row_data = []
                        for col in range(min_col, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            row_data.append(str(cell.value) if cell.value is not None else "")
                        table_data.append("\t".join(row_data))

                    return "\n".join(table_data)
                except Exception as e:
                    return f"Error extracting data table: {str(e)}"

            elif element_type == "comment":
                # Extract comment from cell
                cell_ref = location_data.get("cell", "")

                if not cell_ref or not hasattr(sheet, "comments") or not sheet.comments:
                    return "No comment found"

                if cell_ref in sheet.comments:
                    comment = sheet.comments[cell_ref]
                    author = comment.author if hasattr(comment, 'author') else "Unknown"
                    text = comment.text if hasattr(comment, 'text') else str(comment)
                    return f"Comment by {author}: {text}"
                else:
                    return f"No comment found at cell {cell_ref}"

            elif element_type == "merged_cell":
                # Extract merged cell content
                range_str = location_data.get("range", "")

                if not range_str:
                    return "No range specified for merged cell"

                try:
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(range_str)

                    # Get top-left cell (contains the value for merged cells)
                    cell = sheet.cell(row=min_row, column=min_col)
                    return str(cell.value) if cell.value is not None else ""
                except Exception as e:
                    return f"Error extracting merged cell content: {str(e)}"

            else:
                # Default: return the sheet content as text
                max_row = min(sheet.max_row or 0, self.max_rows)
                max_col = min(sheet.max_column or 0, self.max_cols)

                sheet_content = []
                for row in range(1, max_row + 1):
                    row_values = []
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        row_values.append(str(cell.value) if cell.value is not None else "")
                    sheet_content.append("\t".join(row_values))

                return "\n".join(sheet_content)

        finally:
            # Close workbook
            if wb:
                wb.close()

            # Clean up temporary file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception as e:
                    logger.warning(f"Failed to delete temporary file {temp_file}: {str(e)}")

    def supports_location(self, content_location: str) -> bool:
        """
        Check if this parser supports resolving the given location.

        Args:
            content_location: Content location pointer

        Returns:
            True if supported, False otherwise
        """
        try:
            location_data = json.loads(content_location)
            source = location_data.get("source", "")

            # Check if source exists and is a file
            if not os.path.exists(source) or not os.path.isfile(source):
                return False

            # Check file extension for XLSX
            _, ext = os.path.splitext(source.lower())
            return ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']

        except (json.JSONDecodeError, TypeError):
            return False

    @staticmethod
    def _extract_document_metadata(workbook: openpyxl.Workbook, base_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract metadata from XLSX document.

        Args:
            workbook: The XLSX workbook
            base_metadata: Base metadata from content source

        Returns:
            Dictionary of document metadata
        """
        # Combine base metadata with document properties
        metadata = base_metadata.copy()

        try:
            # Get workbook properties
            props = workbook.properties

            # Add core properties to metadata
            if props.title:
                metadata["title"] = props.title
            if props.creator:
                metadata["author"] = props.creator
            if props.created:
                metadata["created"] = props.created.timestamp() if hasattr(props.created, 'timestamp') else str(
                    props.created)
            if props.modified:
                metadata["modified"] = props.modified.timestamp() if hasattr(props.modified, 'timestamp') else str(
                    props.modified)
            if props.lastModifiedBy:
                metadata["last_modified_by"] = props.lastModifiedBy
            if props.subject:
                metadata["subject"] = props.subject
            if props.keywords:
                metadata["keywords"] = props.keywords
            if props.category:
                metadata["category"] = props.category
            if props.description:
                metadata["description"] = props.description

            # Add document statistics
            metadata["sheet_count"] = len(workbook.sheetnames)
            metadata["sheet_names"] = workbook.sheetnames

            # Add calculation properties if available
            calc_props = {}
            if hasattr(workbook, 'calculation') and workbook.calculation:
                calc = workbook.calculation
                if hasattr(calc, 'calcMode'):
                    calc_props["calc_mode"] = calc.calcMode
                if hasattr(calc, 'calcCompleted'):
                    calc_props["calc_completed"] = calc.calcCompleted
                if hasattr(calc, 'calcOnSave'):
                    calc_props["calc_on_save"] = calc.calcOnSave
                if hasattr(calc, 'fullCalcOnLoad'):
                    calc_props["full_calc_on_load"] = calc.fullCalcOnLoad

            if calc_props:
                metadata["calculation"] = calc_props

        except Exception as e:
            logger.warning(f"Error extracting document metadata: {str(e)}")

        return metadata

    def _parse_workbook(self, workbook: openpyxl.Workbook, doc_id: str, parent_id: str, source_id: str) -> List[
        Dict[str, Any]]:
        """
        Parse Excel workbook into structured elements.

        Args:
            workbook: The openpyxl Workbook
            doc_id: Document ID
            parent_id: Parent element ID
            source_id: Source identifier

        Returns:
            List of parsed elements
        """
        elements = []

        # Create workbook element
        workbook_id = self._generate_id("workbook_")
        workbook_element = {
            "element_id": workbook_id,
            "doc_id": doc_id,
            "element_type": "workbook",
            "parent_id": parent_id,
            "content_preview": f"Excel workbook with {len(workbook.sheetnames)} sheets",
            "content_location": json.dumps({
                "source": source_id,
                "type": "workbook"
            }),
            "content_hash": "",
            "metadata": {
                "sheet_count": len(workbook.sheetnames),
                "active_sheet": workbook.active.title if hasattr(workbook, 'active') and workbook.active else None
            }
        }

        elements.append(workbook_element)

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Skip hidden sheets if not configured to extract them
            if sheet.sheet_state == 'hidden' and not self.extract_hidden_sheets:
                logger.debug(f"Skipping hidden sheet: {sheet_name}")
                continue

            # Process the sheet
            sheet_elements = self._process_sheet(sheet, doc_id, workbook_id, source_id)
            elements.extend(sheet_elements)

        return elements

    def _process_sheet(self, sheet, doc_id: str, parent_id: str, source_id: str) -> List[Dict[str, Any]]:
        """
        Process a single worksheet.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            parent_id: Parent element ID
            source_id: Source identifier

        Returns:
            List of sheet elements
        """
        elements = []

        # Create sheet element
        sheet_id = self._generate_id("sheet_")

        # Get sheet dimensions
        max_row = min(sheet.max_row or 0, self.max_rows)
        max_col = min(sheet.max_column or 0, self.max_cols)

        # Create sheet preview
        if max_row > 0 and max_col > 0:
            preview = f"Sheet '{sheet.title}' with {max_row} rows and {max_col} columns"
        else:
            preview = f"Empty sheet '{sheet.title}'"

        # Sheet metadata
        sheet_metadata = {
            "title": sheet.title,
            "max_row": max_row,
            "max_column": max_col,
            "sheet_state": sheet.sheet_state,  # visible, hidden, or veryHidden
            "tab_color": sheet.sheet_properties.tabColor.rgb if hasattr(sheet.sheet_properties,
                                                                        'tabColor') and sheet.sheet_properties.tabColor else None,
        }

        # Check if sheet has autofilter
        if hasattr(sheet, 'auto_filter') and sheet.auto_filter:
            sheet_metadata["has_autofilter"] = True
            sheet_metadata["autofilter_range"] = str(sheet.auto_filter.ref) if hasattr(sheet.auto_filter,
                                                                                       'ref') else None

        # Check if sheet has freeze panes
        if hasattr(sheet, 'freeze_panes') and sheet.freeze_panes:
            sheet_metadata["has_freeze_panes"] = True
            sheet_metadata["freeze_panes"] = str(sheet.freeze_panes)

        # Create sheet element
        sheet_element = {
            "element_id": sheet_id,
            "doc_id": doc_id,
            "element_type": "sheet",
            "parent_id": parent_id,
            "content_preview": preview,
            "content_location": json.dumps({
                "source": source_id,
                "type": "sheet",
                "sheet_name": sheet.title
            }),
            "content_hash": self._generate_hash(preview),
            "metadata": sheet_metadata
        }

        elements.append(sheet_element)

        # Extract sheet structure
        if max_row > 0 and max_col > 0:
            # Process rows
            row_elements = self._process_rows(sheet, doc_id, sheet_id, source_id, max_row, max_col)
            elements.extend(row_elements)

            # Detect and extract data tables if enabled
            if self.detect_tables and max_row >= self.min_table_rows and max_col >= self.min_table_cols:
                data_table_elements = self._detect_data_tables(sheet, doc_id, sheet_id, source_id, max_row, max_col)
                elements.extend(data_table_elements)

            # Extract merged cells
            merged_cells = self._extract_merged_cells(sheet, doc_id, sheet_id, source_id)
            elements.extend(merged_cells)

            # Extract comments if enabled
            if self.extract_comments:
                comment_elements = self._extract_comments(sheet, doc_id, sheet_id, source_id)
                elements.extend(comment_elements)

        return elements

    def _process_rows(self, sheet, doc_id: str, sheet_id: str, source_id: str, max_row: int, max_col: int) -> List[
        Dict[str, Any]]:
        """
        Process rows from a worksheet.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier
            max_row: Maximum row number
            max_col: Maximum column number

        Returns:
            List of row and cell elements
        """
        elements = []

        # Define column letters for reference
        col_letters = [openpyxl.utils.get_column_letter(col_idx) for col_idx in range(1, max_col + 1)]

        # Process each row
        for row_idx in range(1, max_row + 1):
            # Create row element
            row_id = self._generate_id(f"row_{row_idx}_")

            # Skip entirely empty rows if requested
            # Checking is there is any cell in the row that has a value
            row_has_value = False
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_has_value = True
                    break

            if not row_has_value and self.config.get("skip_empty_rows", True):
                continue

            # Create row element
            row_element = {
                "element_id": row_id,
                "doc_id": doc_id,
                "element_type": "table_row",
                "parent_id": sheet_id,
                "content_preview": f"Row {row_idx}",
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "table_row",
                    "sheet_name": sheet.title,
                    "row": row_idx
                }),
                "content_hash": "",
                "metadata": {
                    "row": row_idx,
                    "sheet": sheet.title,
                    "height": sheet.row_dimensions[row_idx].height if row_idx in sheet.row_dimensions else None,
                    "hidden": sheet.row_dimensions[row_idx].hidden if row_idx in sheet.row_dimensions else False
                }
            }

            elements.append(row_element)

            # Process cells in this row
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)

                # Skip empty cells if requested
                if cell.value is None and self.config.get("skip_empty_cells", True):
                    continue

                # Get cell address (e.g., A1, B2)
                cell_addr = f"{col_letters[col_idx - 1]}{row_idx}"

                # Create cell element
                cell_id = self._generate_id(f"cell_{cell_addr}_")

                # Format cell value for display
                cell_value = cell.value
                if cell_value is not None:
                    if isinstance(cell_value, (int, float)):
                        content_preview = str(cell_value)
                    else:
                        content_preview = str(cell_value)
                else:
                    content_preview = ""

                # Limit preview length
                if len(content_preview) > self.max_content_preview:
                    content_preview = content_preview[:self.max_content_preview - 3] + "..."

                # Cell metadata
                cell_metadata = {
                    "address": cell_addr,
                    "row": row_idx,
                    "column": col_idx,
                    "column_letter": col_letters[col_idx - 1],
                    "data_type": cell.data_type,
                    "style": self._extract_cell_style(cell)
                }

                # Add formula if available and extracting formulas is enabled
                if self.extract_formulas and cell.formula:
                    cell_metadata["formula"] = cell.formula

                # Create cell element
                cell_element = {
                    "element_id": cell_id,
                    "doc_id": doc_id,
                    "element_type": "table_cell",
                    "parent_id": row_id,
                    "content_preview": content_preview,
                    "content_location": json.dumps({
                        "source": source_id,
                        "type": "table_cell",
                        "sheet_name": sheet.title,
                        "cell": cell_addr
                    }),
                    "content_hash": self._generate_hash(content_preview),
                    "metadata": cell_metadata
                }

                # If this is a header row (row 1), mark as table_header
                if row_idx == 1:
                    cell_element["element_type"] = "table_header"

                elements.append(cell_element)

        return elements

    def _extract_merged_cells(self, sheet, doc_id: str, sheet_id: str, source_id: str) -> List[Dict[str, Any]]:
        """
        Extract merged cells information.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier

        Returns:
            List of merged cell elements
        """
        elements = []

        if not hasattr(sheet, 'merged_cells') or not sheet.merged_cells:
            return elements

        for merged_range in sheet.merged_cells.ranges:
            # Create merged cell element
            merged_id = self._generate_id("merged_")

            # Get coordinate information
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

            # Get the value from the top-left cell of the merge
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            value = top_left_cell.value
            content_preview = str(value) if value is not None else ""

            # Limit preview length
            if len(content_preview) > self.max_content_preview:
                content_preview = content_preview[:self.max_content_preview - 3] + "..."

            # Create merged cell element
            merged_element = {
                "element_id": merged_id,
                "doc_id": doc_id,
                "element_type": "merged_cell",
                "parent_id": sheet_id,
                "content_preview": content_preview,
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "merged_cell",
                    "sheet_name": sheet.title,
                    "range": str(merged_range)
                }),
                "content_hash": self._generate_hash(content_preview),
                "metadata": {
                    "range": str(merged_range),
                    "min_row": min_row,
                    "min_column": min_col,
                    "max_row": max_row,
                    "max_column": max_col,
                    "sheet": sheet.title
                }
            }

            elements.append(merged_element)

        return elements

    def _extract_comments(self, sheet, doc_id: str, sheet_id: str, source_id: str) -> List[Dict[str, Any]]:
        """
        Extract comments from worksheet.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier

        Returns:
            List of comment elements
        """
        elements = []

        # Check if there are comments
        if not hasattr(sheet, 'comments') or not sheet.comments:
            return elements

        # Process comments
        for cell_addr, comment in sheet.comments.items():
            # Create comment element
            comment_id = self._generate_id("comment_")

            # Extract comment text and author
            text = comment.text if hasattr(comment, 'text') else str(comment)
            author = comment.author if hasattr(comment, 'author') else "Unknown"

            # Limit text length for preview
            content_preview = f"Comment by {author}: {text}"
            if len(content_preview) > self.max_content_preview:
                content_preview = content_preview[:self.max_content_preview - 3] + "..."

            # Create comment element
            comment_element = {
                "element_id": comment_id,
                "doc_id": doc_id,
                "element_type": "comment",
                "parent_id": sheet_id,
                "content_preview": content_preview,
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "comment",
                    "sheet_name": sheet.title,
                    "cell": cell_addr
                }),
                "content_hash": self._generate_hash(text),
                "metadata": {
                    "cell": cell_addr,
                    "author": author,
                    "text": text,
                    "sheet": sheet.title
                }
            }

            elements.append(comment_element)

        return elements

    @staticmethod
    def _extract_cell_style(cell) -> Dict[str, Any]:
        """
        Extract style information from cell.

        Args:
            cell: The cell object

        Returns:
            Dictionary with style information
        """
        style = {}

        # Extract basic style properties
        if hasattr(cell, 'font') and cell.font:
            font = {}
            if hasattr(cell.font, 'bold') and cell.font.bold:
                font["bold"] = True
            if hasattr(cell.font, 'italic') and cell.font.italic:
                font["italic"] = True
            if hasattr(cell.font, 'underline') and cell.font.underline:
                font["underline"] = True
            if hasattr(cell.font, 'strike') and cell.font.strike:
                font["strike"] = True
            if hasattr(cell.font, 'color') and cell.font.color:
                font["color"] = cell.font.color.rgb if hasattr(cell.font.color, 'rgb') else None
            if hasattr(cell.font, 'name') and cell.font.name:
                font["name"] = cell.font.name
            if hasattr(cell.font, 'size') and cell.font.size:
                font["size"] = cell.font.size

            if font:
                style["font"] = font

        # Extract alignment
        if hasattr(cell, 'alignment') and cell.alignment:
            alignment = {}
            if hasattr(cell.alignment, 'horizontal') and cell.alignment.horizontal:
                alignment["horizontal"] = cell.alignment.horizontal
            if hasattr(cell.alignment, 'vertical') and cell.alignment.vertical:
                alignment["vertical"] = cell.alignment.vertical
            if hasattr(cell.alignment, 'wrap_text') and cell.alignment.wrap_text:
                alignment["wrap_text"] = True

            if alignment:
                style["alignment"] = alignment

        # Extract fill
        if hasattr(cell, 'fill') and cell.fill:
            fill = {}
            if hasattr(cell.fill, 'fill_type') and cell.fill.fill_type:
                fill["type"] = cell.fill.fill_type
            if hasattr(cell.fill, 'start_color') and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                fill["color"] = cell.fill.start_color.rgb

            if fill:
                style["fill"] = fill

        # Extract border
        if hasattr(cell, 'border') and cell.border:
            border = {}
            for side in ['left', 'right', 'top', 'bottom']:
                side_border = getattr(cell.border, side, None)
                if side_border and hasattr(side_border, 'style') and side_border.style:
                    if 'sides' not in border:
                        border["sides"] = {}
                    border["sides"][side] = {
                        "style": side_border.style,
                        "color": side_border.color.rgb if hasattr(side_border,
                                                                  'color') and side_border.color and hasattr(
                            side_border.color, 'rgb') else None
                    }

            if border:
                style["border"] = border

        # Extract number format
        if hasattr(cell, 'number_format') and cell.number_format:
            style["number_format"] = cell.number_format

        return style

    def _detect_data_tables(self, sheet, doc_id: str, sheet_id: str, source_id: str, max_row: int, max_col: int) -> \
            List[Dict[str, Any]]:
        """
        Detect and extract structured data tables within a worksheet.
        This focuses on finding regions that appear to be 2D tables with headers.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier
            max_row: Maximum row number
            max_col: Maximum column number

        Returns:
            List of data table elements
        """
        elements = []

        # Skip small sheets that are unlikely to contain meaningful tables
        if max_row < self.min_table_rows or max_col < self.min_table_cols:
            return elements

        # Use heuristics to detect potential tables
        # 1. Check for consistent data regions with header rows
        # 2. Look for formatting patterns (e.g., header row formatting)
        # 3. Check for the presence of autofilters (strong indicator of tables)

        # Start by detecting candidate table regions
        table_regions = []

        # If an autofilter is present, it's a strong indication of a table
        if hasattr(sheet, 'auto_filter') and sheet.auto_filter and hasattr(sheet.auto_filter, 'ref'):
            autofilter_range = sheet.auto_filter.ref
            if autofilter_range:
                # Parse the range (e.g., "A1:F20")
                try:
                    # Get the coordinates from the range
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(autofilter_range)

                    table_regions.append({
                        "min_row": min_row,
                        "min_col": min_col,
                        "max_row": max_row,
                        "max_col": max_col,
                        "has_header": True,  # Assume first row is header in autofilter
                        "confidence": "high"  # High confidence due to autofilter
                    })
                except Exception as e:
                    logger.debug(f"Error parsing autofilter range: {str(e)}")

        # If no autofilter, try to detect tables by analyzing data patterns
        if not table_regions:
            # Get a snapshot of the data
            data_snapshot = []
            for row_idx in range(1, min(max_row + 1, 20)):  # Limit to first 20 rows for performance
                row_data = []
                empty_count = 0
                for col_idx in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns for performance
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    is_empty = cell.value is None
                    row_data.append({
                        "value": cell.value,
                        "is_empty": is_empty,
                        "is_bold": hasattr(cell, 'font') and cell.font and hasattr(cell.font,
                                                                                   'bold') and cell.font.bold,
                        "has_fill": hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill, 'fill_type') and
                                    cell.fill.fill_type != 'none'
                    })
                    if is_empty:
                        empty_count += 1

                # Skip completely empty rows
                if empty_count == len(row_data):
                    continue

                data_snapshot.append(row_data)

            # Simple table detection: Look for patterns like:
            # - First row has formatting different from other rows (likely headers)
            # - Consistent data in columns
            # - Few empty cells in the middle of data

            if len(data_snapshot) >= 2:  # Need at least 2 rows (header + data)
                first_row = data_snapshot[0]

                # Check if first row might be a header row
                header_indicators = 0
                for cell_data in first_row:
                    if cell_data["is_bold"] or cell_data["has_fill"]:
                        header_indicators += 1

                # If more than half of cells in first row have header-like formatting
                is_likely_header = header_indicators > len(first_row) / 2

                # Find how many columns might be in the table
                max_col_with_data = 0
                for row_data in data_snapshot:
                    for col_idx, cell_data in enumerate(row_data):
                        if not cell_data["is_empty"]:
                            max_col_with_data = max(max_col_with_data, col_idx + 1)

                # Check for consistent data patterns in columns
                if max_col_with_data >= 2:  # Need at least 2 columns
                    # Calculate approximate table boundaries
                    table_min_row = 1
                    table_min_col = 1

                    # Determine table height: look for empty rows or significant formatting changes
                    table_max_row = len(data_snapshot)
                    for row_idx in range(1, len(data_snapshot)):
                        empty_count = sum(1 for cell in data_snapshot[row_idx] if cell["is_empty"])
                        if empty_count > max_col_with_data / 2:  # Over half the cells are empty
                            # This might be the end of the table
                            table_max_row = row_idx
                            break

                    table_max_col = max_col_with_data

                    # If we found a viable table region, add it
                    if table_max_row >= 2 and table_max_col >= 2:
                        table_regions.append({
                            "min_row": table_min_row,
                            "min_col": table_min_col,
                            "max_row": table_max_row,
                            "max_col": table_max_col,
                            "has_header": is_likely_header,
                            "confidence": "medium"  # Medium confidence due to heuristics
                        })

        # Process each detected table region
        for idx, region in enumerate(table_regions):
            # Create a data table element
            table_id = self._generate_id(f"data_table_{idx + 1}_")

            # Generate a table preview with header row if present
            preview = f"Data table {idx + 1} ({region['max_row'] - region['min_row'] + 1}x{region['max_col'] - region['min_col'] + 1})"
            if region["has_header"]:
                header_row_idx = region["min_row"]
                header_values = []
                for col_idx in range(region["min_col"], region["max_col"] + 1):
                    cell = sheet.cell(row=header_row_idx, column=col_idx)
                    if cell.value is not None:
                        header_values.append(str(cell.value))

                if header_values:
                    preview += f" with headers: {', '.join(header_values[:3])}"
                    if len(header_values) > 3:
                        preview += ", ..."

            # Generate a textual representation of the table for searching
            table_content = []
            for row_idx in range(region["min_row"], region["max_row"] + 1):
                row_values = []
                for col_idx in range(region["min_col"], region["max_col"] + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    row_values.append(str(cell.value) if cell.value is not None else "")
                table_content.append("\t".join(row_values))

            # Join all rows with newlines to create a searchable text representation
            table_text = "\n".join(table_content)

            # Get range in Excel notation (e.g., "A1:F20")
            from openpyxl.utils import get_column_letter
            min_col_letter = get_column_letter(region["min_col"])
            max_col_letter = get_column_letter(region["max_col"])
            range_str = f"{min_col_letter}{region['min_row']}:{max_col_letter}{region['max_row']}"

            # Create data table element
            table_element = {
                "element_id": table_id,
                "doc_id": doc_id,
                "element_type": "data_table",
                "parent_id": sheet_id,
                "content_preview": preview,
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "data_table",
                    "sheet_name": sheet.title,
                    "range": range_str
                }),
                "content_hash": self._generate_hash(table_text),
                "metadata": {
                    "range": range_str,
                    "min_row": region["min_row"],
                    "min_col": region["min_col"],
                    "max_row": region["max_row"],
                    "max_col": region["max_col"],
                    "row_count": region["max_row"] - region["min_row"] + 1,
                    "column_count": region["max_col"] - region["min_col"] + 1,
                    "has_header": region["has_header"],
                    "detection_confidence": region["confidence"],
                    "sheet": sheet.title,
                    "table_contents": table_text  # Include text version for searching
                }
            }

            elements.append(table_element)

            # Add separate elements for header row and first column if they're likely headers
            if region["has_header"]:
                # Extract header row
                header_id = self._generate_id(f"table_header_row_{idx + 1}_")
                header_values = []
                for col_idx in range(region["min_col"], region["max_col"] + 1):
                    cell = sheet.cell(row=region["min_row"], column=col_idx)
                    header_values.append(str(cell.value) if cell.value is not None else "")

                header_text = "\t".join(header_values)

                # Create header row element
                header_element = {
                    "element_id": header_id,
                    "doc_id": doc_id,
                    "element_type": "table_header_row",
                    "parent_id": table_id,
                    "content_preview": header_text[:self.max_content_preview] + (
                        "..." if len(header_text) > self.max_content_preview else ""),
                    "content_location": json.dumps({
                        "source": source_id,
                        "type": "table_header_row",
                        "sheet_name": sheet.title,
                        "range": f"{min_col_letter}{region['min_row']}:{max_col_letter}{region['min_row']}"
                    }),
                    "content_hash": self._generate_hash(header_text),
                    "metadata": {
                        "row": region["min_row"],
                        "values": header_values,
                        "sheet": sheet.title
                    }
                }

                elements.append(header_element)

                # Check if first column might contain row headers
                # Simple heuristic: Check if formatting of first column is different
                first_col_headers = []
                first_col_formatting_count = 0

                for row_idx in range(region["min_row"] + 1, region["max_row"] + 1):  # Skip header row
                    cell = sheet.cell(row=row_idx, column=region["min_col"])
                    first_col_headers.append(str(cell.value) if cell.value is not None else "")

                    # Check if the cell has special formatting (bold, fill, etc.)
                    if (hasattr(cell, 'font') and cell.font and hasattr(cell.font, 'bold') and cell.font.bold) or \
                            (hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill,
                                                                             'fill_type') and cell.fill.fill_type != 'none'):
                        first_col_formatting_count += 1

                # If more than 1/3 of cells in the first column have special formatting, consider it a header column
                if first_col_formatting_count > (region["max_row"] - region["min_row"]) / 3:
                    col_header_id = self._generate_id(f"table_row_headers_{idx + 1}_")
                    col_header_text = "\n".join(first_col_headers)

                    # Create row headers element
                    col_header_element = {
                        "element_id": col_header_id,
                        "doc_id": doc_id,
                        "element_type": "table_row_headers",
                        "parent_id": table_id,
                        "content_preview": col_header_text[:self.max_content_preview] + (
                            "..." if len(col_header_text) > self.max_content_preview else ""),
                        "content_location": json.dumps({
                            "source": source_id,
                            "type": "table_row_headers",
                            "sheet_name": sheet.title,
                            "range": f"{min_col_letter}{region['min_row'] + 1}:{min_col_letter}{region['max_row']}"
                        }),
                        "content_hash": self._generate_hash(col_header_text),
                        "metadata": {
                            "column": region["min_col"],
                            "values": first_col_headers,
                            "sheet": sheet.title
                        }
                    }

                    elements.append(col_header_element)

        return elements

    def _extract_links(self, content: str, element_id: str) -> List[Dict[str, Any]]:
        """
        Extract links from content.

        Args:
            content: Text content
            element_id: ID of the element containing the links

        Returns:
            List of extracted link dictionaries
        """
        links = []

        # For Excel, we need to handle links differently than in text-based documents
        # Since we can't extract hyperlinks directly from content string
        # Instead, we'll implement a helper method for extracting links during the parsing phase

        # This base implementation returns an empty list
        # The actual link extraction happens during the parsing phase in _extract_workbook_links

        return links

    @staticmethod
    def _extract_workbook_links(elements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Helper method to extract hyperlinks from workbook elements.
        This is called during the parsing phase.

        Args:
            elements: List of extracted elements

        Returns:
            List of hyperlink dictionaries
        """
        links = []

        # Find cells with hyperlinks in the metadata
        for element in elements:
            if element["element_type"] == "table_cell" and "metadata" in element:
                metadata = element["metadata"]

                # Check if hyperlink is in the cell style
                if "style" in metadata and "hyperlink" in metadata["style"]:
                    hyperlink = metadata["style"]["hyperlink"]

                    # Extract hyperlink details
                    link_target = hyperlink.get("target", "")
                    link_text = hyperlink.get("display", element.get("content_preview", ""))

                    if link_target:
                        links.append({
                            "source_id": element["element_id"],
                            "link_text": link_text,
                            "link_target": link_target,
                            "link_type": "hyperlink"
                        })

        return links

    @staticmethod
    def _generate_hash(content: str) -> str:
        """
        Generate a hash of content for change detection.

        Args:
            content: Text content

        Returns:
            MD5 hash of content
        """
        return hashlib.md5(content.encode('utf-8')).hexdigest()
